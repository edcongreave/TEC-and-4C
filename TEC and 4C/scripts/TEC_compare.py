import os
import glob
import re
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- CONSTANTS ---
DATA_FOLDER = "TEC files"
KEY_COLUMN = 'Project ID'
CAPACITY_COLUMN = 'Cumulative Total Capacity (MW)'
PLANT_TYPE_COLUMN = 'Plant Type'
DATE_COLUMN = 'MW Effective From'

# Columns to track for changes (and display in the detailed report)
# These are the DESIRED columns - the code will handle if they're missing
COMPARISON_COLUMNS = [
    'Project Name', 'Customer Name', 'Connection Site', 'Stage', 'MW Connected', 
    'MW Increase / Decrease', CAPACITY_COLUMN, DATE_COLUMN, 'Project Status', 
    'Agreement Type', 'HOST TO', PLANT_TYPE_COLUMN, 'Project Number', 'Gate'
]

# --- Helper Functions (File Selection and Loading) ---

def get_file_list(folder_path=DATA_FOLDER):
    """Scans the specified folder for CSV files."""
    if not os.path.isdir(folder_path):
        print(f"\nError: The data folder '{folder_path}' does not exist.")
        return {}, []
    search_pattern = os.path.join(folder_path, "*.csv")
    all_csv_files = sorted(glob.glob(search_pattern))
    date_pattern = re.compile(r'tec[_-]register[_-]?(\d{4}-\d{2}-\d{2})', re.IGNORECASE)
    files_by_date = {}
    valid_files = []
    for full_path in all_csv_files:
        filename = os.path.basename(full_path)
        match = date_pattern.search(filename)
        if match:
            date_str = match.group(1)
            files_by_date.setdefault(date_str, []).append(filename)
            valid_files.append(full_path)
    if not valid_files:
        print(f"\nError: No TEC register CSV files found in '{folder_path}' matching the naming pattern.")
    return files_by_date, valid_files

def get_comparison_selection(files_by_date, valid_files):
    """Prompts the user to select two files for comparison."""
    if not valid_files:
        return None, None
    print("\n--- Available Data Versions ---")
    indexed_files_map = {}
    current_index = 1
    sorted_dates = sorted(files_by_date.keys())
    for date_str in sorted_dates:
        print(f"\n[{date_str}]")
        for filename in sorted(files_by_date[date_str]):
            full_path = os.path.join(DATA_FOLDER, filename)
            indexed_files_map[current_index] = full_path
            print(f"  {current_index}: {filename}")
            current_index += 1
    print("-" * 35)
    def select_file(prompt):
        while True:
            try:
                choice = input(prompt)
                if not choice: return None
                index = int(choice)
                if index in indexed_files_map:
                    return indexed_files_map.get(index)
                else:
                    print(f"Invalid number. Please select a number between 1 and {len(indexed_files_map)}.")
            except ValueError:
                print("Invalid input. Please enter a number.")
    file_a_path = select_file("Select the first file (Version A, Older) by its number: ")
    if not file_a_path: return None, None
    file_b_path = select_file("Select the second file (Version B, Newer) by its number (Press ENTER to skip comparison): ")
    if file_b_path is None: file_b_path = file_a_path 
    return file_a_path, file_b_path

def ensure_missing_columns(df, required_columns):
    """
    Ensures all required columns exist in the DataFrame.
    Adds missing columns with empty/null values.
    """
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''  # Add missing column with empty string
            print(f"  Note: Column '{col}' not found - added with blank values")
    return df

def load_raw_data(file_path):
    """
    Loads a CSV, cleans columns, handles date formats, and adds a Pot category.
    Now robust to missing columns.
    """
    try:
        df = pd.read_csv(file_path, low_memory=False)
        df.columns = df.columns.str.strip()
        
        df.drop(columns=['_id'], errors='ignore', inplace=True)
        
        if CAPACITY_COLUMN not in df.columns and len(df.columns) > 6:
            df.rename(columns={df.columns[6]: CAPACITY_COLUMN}, inplace=True)
        elif CAPACITY_COLUMN not in df.columns:
            raise KeyError(f"Required column '{CAPACITY_COLUMN}' not found.")
        
        df[KEY_COLUMN] = df[KEY_COLUMN].astype(str).str.strip()
        df.dropna(subset=[KEY_COLUMN], inplace=True)
        
        # --- FIX: Smarter Date Parsing to avoid warnings ---
        if DATE_COLUMN in df.columns:
            uses_slash_format = df[DATE_COLUMN].astype(str).str.contains('/').any()

            if uses_slash_format:
                df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN], errors='coerce', dayfirst=True).dt.normalize()
            else:
                df[DATE_COLUMN] = pd.to_datetime(df[DATE_COLUMN], errors='coerce').dt.normalize()
        
        df[CAPACITY_COLUMN] = pd.to_numeric(df[CAPACITY_COLUMN], errors='coerce').round(2)
        
        # Handle missing Plant Type column
        if PLANT_TYPE_COLUMN in df.columns:
            df['Pot'] = df[PLANT_TYPE_COLUMN].apply(
                lambda x: 'Pot 3' if str(x).lower().strip() in ['wind offshore', 'offshore wind'] else 
                          ('Pot 1' if str(x).lower().strip() in ['wind onshore', 'onshore wind', 'pv array (photo voltaic/solar)', 'solar'] else 'Other')
            )
        else:
            df['Pot'] = 'Other'
            print(f"  Note: Column '{PLANT_TYPE_COLUMN}' not found - all projects categorized as 'Other'")
        
        # Ensure all comparison columns exist
        df = ensure_missing_columns(df, COMPARISON_COLUMNS)
        
        return df
    
    except Exception as e:
        print(f"Error loading {os.path.basename(file_path)}: {e}")
        return None

# --- Core Comparison Logic (Final Version) ---

def prepare_df_for_report(df, cols_to_keep):
    """
    Prepares a change-group DataFrame for final concatenation and handles empty inputs.
    """
    final_df = df.copy()
    
    # Define the necessary output columns to maintain structure for concatenation
    report_cols_all = COMPARISON_COLUMNS + ['Pot']
    final_output_cols = [KEY_COLUMN, 'Project Name', 'Changes_Description', 'Change_Type'] + [c for c in report_cols_all if c not in [KEY_COLUMN, 'Project Name']]
    
    # --- FIX: Check for empty DataFrame to avoid IndexError ---
    if final_df.empty:
        return pd.DataFrame(columns=final_output_cols) 

    # Rename A columns in the REMOVED DF to match the output structure
    if 'Change_Type' in final_df.columns and final_df['Change_Type'].iloc[0] == 'REMOVED':
         # This block is for REMOVED data (A data)
         final_df = final_df.rename(columns={col: col for col in cols_to_keep}, errors='ignore').copy()
    
    # Filter the DataFrame to the required output columns (only those that exist)
    available_cols = [col for col in final_output_cols if col in final_df.columns]
    return final_df[available_cols]


def get_project_level_comparison(df_a, df_b):
    """
    1. Aggregates data to the Project ID level to ensure unique comparison and reduce noise.
    2. Performs a field-by-field comparison, generating a descriptive string for changes.
    Now robust to missing columns.
    """
    
    def aggregate_for_comparison(df):
        """Groups data by Project ID for a unique comparison."""
        # Only aggregate columns that actually exist in the DataFrame
        agg_dict = {}
        for col in COMPARISON_COLUMNS:
            if col in df.columns:
                if col == CAPACITY_COLUMN:
                    agg_dict[col] = 'sum'
                else:
                    agg_dict[col] = 'first'
        
        if 'Pot' in df.columns:
            agg_dict['Pot'] = 'first'
        
        return df.groupby(KEY_COLUMN)[list(agg_dict.keys())].agg(agg_dict)

    # 1. Aggregate and Merge
    df_a_agg = aggregate_for_comparison(df_a).reset_index()
    df_b_agg = aggregate_for_comparison(df_b).reset_index()
    
    df_merged_status = df_a_agg.merge(
        df_b_agg, on=KEY_COLUMN, suffixes=('_A', '_B'), how='outer', indicator=True
    )
    
    added_ids = df_merged_status[df_merged_status['_merge'] == 'right_only'][KEY_COLUMN].tolist()
    removed_ids = df_merged_status[df_merged_status['_merge'] == 'left_only'][KEY_COLUMN].tolist()
    
    df_common = df_a_agg.merge(
        df_b_agg, on=KEY_COLUMN, suffixes=('_A', '_B'), how='inner'
    )

    # 2. Detailed Field-Level Comparison for common projects
    change_descriptions = []
    
    for index, row in df_common.iterrows():
        descriptions = []
        for col in COMPARISON_COLUMNS:
            col_a = col + '_A'
            col_b = col + '_B'
            
            # Skip if column doesn't exist in either version
            if col_a not in row.index or col_b not in row.index:
                continue
            
            val_a_raw = row[col_a]
            val_b_raw = row[col_b]

            changed = False
            
            # Helper to format values for display
            val_a_display = str(val_a_raw) if pd.notna(val_a_raw) else 'NULL'
            val_b_display = str(val_b_raw) if pd.notna(val_b_raw) else 'NULL'

            if col == DATE_COLUMN:
                # Comparison: normalize to string for comparison
                val_a_str = val_a_raw.strftime('%Y-%m-%d') if pd.notna(val_a_raw) else 'NULL'
                val_b_str = val_b_raw.strftime('%Y-%m-%d') if pd.notna(val_b_raw) else 'NULL'
                
                changed = (val_a_str != val_b_str)
                # Display: use D/M/Y format
                val_a_display = val_a_raw.strftime('%d/%m/%Y') if pd.notna(val_a_raw) else 'NULL'
                val_b_display = val_b_raw.strftime('%d/%m/%Y') if pd.notna(val_b_raw) else 'NULL'
                
            elif col == CAPACITY_COLUMN:
                # Comparison: Numerics with tolerance
                val_a_num = val_a_raw if pd.notna(val_a_raw) else 0.0
                val_b_num = val_b_raw if pd.notna(val_b_raw) else 0.0
                
                changed = abs(val_a_num - val_b_num) > 0.01
                # Display: formatted capacity
                val_a_display = f"{val_a_num:,.2f} MW"
                val_b_display = f"{val_b_num:,.2f} MW"
                
            else:
                # Comparison: Strings (strip whitespace, handle NULLs)
                val_a_str = str(val_a_raw).strip() if pd.notna(val_a_raw) else 'NULL'
                val_b_str = str(val_b_raw).strip() if pd.notna(val_b_raw) else 'NULL'
                
                # Special handling for Gate column: ignore changes to blank/NULL
                if col == 'Gate':
                    # Only report as changed if the new value is meaningful (not blank/NULL)
                    if val_b_str in ['', 'NULL', 'nan']:
                        changed = False
                    else:
                        changed = (val_a_str != val_b_str)
                # Special handling for Project Number: ignore trailing -0 differences
                elif col == 'Project Number':
                    # Normalize by removing trailing -0
                    val_a_normalized = val_a_str.rstrip('-0') if val_a_str != 'NULL' else 'NULL'
                    val_b_normalized = val_b_str.rstrip('-0') if val_b_str != 'NULL' else 'NULL'
                    changed = (val_a_normalized != val_b_normalized)
                # Special handling for Stage: ignore formatting differences and 0/NULL equivalence
                elif col == 'Stage':
                    # Normalize stage values: convert to float, treat 0 as NULL
                    def normalize_stage(val_str):
                        if val_str in ['', 'NULL', 'nan']:
                            return 'NULL'
                        try:
                            # Try to convert to float to normalize (e.g., "1.0" -> 1.0)
                            stage_float = float(val_str)
                            # Treat 0 as NULL
                            if stage_float == 0.0:
                                return 'NULL'
                            # Return as integer if it's a whole number
                            return str(int(stage_float)) if stage_float.is_integer() else str(stage_float)
                        except (ValueError, AttributeError):
                            return val_str
                    
                    val_a_normalized = normalize_stage(val_a_str)
                    val_b_normalized = normalize_stage(val_b_str)
                    changed = (val_a_normalized != val_b_normalized)
                else:
                    changed = (val_a_str != val_b_str)
                
                # Display: clean strings
                val_a_display = val_a_str
                val_b_display = val_b_str
            
            if changed:
                descriptions.append(f"{col}: {val_a_display} -> {val_b_display}")

        change_descriptions.append("; ".join(descriptions))
    
    df_common['Changes_Description'] = change_descriptions
    
    df_changed_agg = df_common[df_common['Changes_Description'] != ''].copy()
    
    # 3. Final Report Assembly (uses full raw data from A or B)
    report_cols = [col for col in COMPARISON_COLUMNS + ['Pot'] if col in df_a.columns or col in df_b.columns]

    df_added = df_b[df_b[KEY_COLUMN].isin(added_ids)].copy()
    df_added['Change_Type'] = 'ADDED'
    df_added['Changes_Description'] = 'NEWLY ADDED PROJECT'
    
    df_removed = df_a[df_a[KEY_COLUMN].isin(removed_ids)].copy()
    df_removed['Change_Type'] = 'REMOVED'
    df_removed['Changes_Description'] = 'REMOVED PROJECT'
    
    df_changed_latest = df_b[df_b[KEY_COLUMN].isin(df_changed_agg[KEY_COLUMN])].copy()
    
    df_changed_latest = df_changed_latest.merge(
        df_changed_agg[[KEY_COLUMN, 'Changes_Description']],
        on=KEY_COLUMN,
        how='left'
    )
    df_changed_latest['Change_Type'] = 'CHANGED'

    # Prepare each group, using the fixed function
    df_added_report = prepare_df_for_report(df_added, report_cols)
    df_removed_report = prepare_df_for_report(df_removed, report_cols)
    df_changed_report = prepare_df_for_report(df_changed_latest, report_cols)

    # Concatenate the three groups (handles empty DFs safely now)
    df_final_report = pd.concat([
        df_added_report,
        df_removed_report,
        df_changed_report
    ], ignore_index=True)

    # --- Summary Report Generation ---
    pot_col = 'Pot_B' if 'Pot_B' in df_merged_status.columns else 'Pot_A'
    df_change_summary_report = df_merged_status[[KEY_COLUMN, pot_col, '_merge']].copy()
    df_change_summary_report.rename(columns={pot_col: 'Pot', '_merge': 'Change_Type'}, inplace=True)
    
    df_change_summary_report['Change_Type'] = df_change_summary_report['Change_Type'].astype(object)
    
    df_change_summary_report['Change_Type'] = df_change_summary_report['Change_Type'].replace({
        'right_only': 'ADDED', 
        'left_only': 'REMOVED', 
        'both': 'UNCHANGED'
    })
    
    changed_ids_only = df_changed_agg[KEY_COLUMN].tolist()
    df_change_summary_report.loc[df_change_summary_report[KEY_COLUMN].isin(changed_ids_only), 'Change_Type'] = 'CHANGED'
    
    # Merge capacity data for summary
    if CAPACITY_COLUMN in df_b_agg.columns:
        df_capacity_b = df_b_agg.reset_index()[[KEY_COLUMN, CAPACITY_COLUMN]].rename(columns={CAPACITY_COLUMN: 'Capacity'})
        df_change_summary_report = df_change_summary_report.merge(df_capacity_b, on=KEY_COLUMN, how='left')
    
    if CAPACITY_COLUMN in df_a_agg.columns:
        df_capacity_a = df_a_agg.reset_index()[[KEY_COLUMN, CAPACITY_COLUMN]].rename(columns={CAPACITY_COLUMN: 'Capacity_A'})
        df_change_summary_report = df_change_summary_report.merge(df_capacity_a, on=KEY_COLUMN, how='left')
        
        if 'Capacity' in df_change_summary_report.columns:
            df_change_summary_report['Capacity'] = df_change_summary_report['Capacity'].fillna(df_change_summary_report['Capacity_A'])
            df_change_summary_report.drop(columns=['Capacity_A'], inplace=True)
        else:
            df_change_summary_report.rename(columns={'Capacity_A': 'Capacity'}, inplace=True)

    return df_change_summary_report, df_final_report

def create_summary_df(change_data, pot_name):
    """Creates a summary DataFrame for a given Pot using aggregated project data."""
    df = change_data[change_data['Pot'].fillna('Other') == pot_name]
    
    added_capacity = df[df['Change_Type'] == 'ADDED']['Capacity'].sum().round(2) if 'Capacity' in df.columns else 0
    removed_capacity = df[df['Change_Type'] == 'REMOVED']['Capacity'].sum().round(2) if 'Capacity' in df.columns else 0
    
    summary = {
        'Metric': [
            'Projects Added (in B only)',
            'Projects Removed (in A only)',
            'Projects with Field Change (excluding Added/Removed)',
            '---',
            'Capacity Added (MW)',
            'Capacity Removed (MW)',
        ],
        'Value': [
            len(df[df['Change_Type'] == 'ADDED']),
            len(df[df['Change_Type'] == 'REMOVED']),
            len(df[df['Change_Type'] == 'CHANGED']),
            '---',
            f"{added_capacity:,.2f}",
            f"{removed_capacity:,.2f}"
        ]
    }
    return pd.DataFrame(summary)

def write_to_excel(df_change_summary, df_final_report, file_a_path, file_b_path):
    """Writes the comparison results to a multi-sheet Excel file with custom formatting."""
    
    version_b_name = os.path.basename(file_b_path).replace('.csv', '').replace('tec-register-', '')
    output_filename = f"TEC_Comparison_Report_DESCRIPTIVE_{version_b_name}.xlsx"
    
    pot3_name = "Pot 3 (Offshore Wind)"
    pot1_name = "Pot 1 (Onshore-Solar)" 
    
    pot3_report = df_final_report[df_final_report['Pot'] == 'Pot 3'].copy()
    if 'Pot' in pot3_report.columns:
        pot3_report = pot3_report.drop(columns=['Pot'])
    
    pot1_report = df_final_report[df_final_report['Pot'] == 'Pot 1'].copy()
    if 'Pot' in pot1_report.columns:
        pot1_report = pot1_report.drop(columns=['Pot'])

    pot3_summary = create_summary_df(df_change_summary, 'Pot 3')
    pot1_summary = create_summary_df(df_change_summary, 'Pot 1')

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            
            if "Sheet" in writer.book.sheetnames:
                del writer.book["Sheet"]

            sheets_data = {
                pot3_name: (pot3_report, pot3_summary, file_a_path, file_b_path),
                pot1_name: (pot1_report, pot1_summary, file_a_path, file_b_path),
            }

            for sheet_name, (detailed_df, summary_df, file_a_name, file_b_name) in sheets_data.items():
                
                ws = writer.book.create_sheet(sheet_name)
                
                # --- Write Header & Summary ---
                ws.append([f"TEC Register Comparison: {sheet_name}"])
                ws.append([f"Version A: {os.path.basename(file_a_name)}", f"Version B: {os.path.basename(file_b_name)}"])
                ws.append([])
                
                ws.append(["SUMMARY OF PROJECT CHANGES:"])
                ws.append([]) 
                
                start_row_summary = ws.max_row + 1
                for r in dataframe_to_rows(summary_df, header=True, index=False):
                    ws.append(r)
                
                summary_header_row = ws[start_row_summary]
                for cell in summary_header_row:
                    cell.font = Font(bold=True)
                
                ws.append([])
                ws.append(["--- DETAILED PROJECT CHANGES ---"])
                ws.append(["'ADDED' and 'CHANGED' rows show Version B (Latest) data."])
                ws.append(["'REMOVED' rows show Version A (Old) data."])
                ws.append([])

                # --- Write Detailed Changes (Version B data + Descriptive Changes) ---
                start_row_detailed = ws.max_row + 1

                desired_order = [KEY_COLUMN, 'Project Name', 'Changes_Description', 'Change_Type']
                remaining_cols = [c for c in detailed_df.columns if c not in desired_order]
                display_cols = desired_order + remaining_cols
                
                # Only use columns that exist in the DataFrame
                display_cols = [col for col in display_cols if col in detailed_df.columns]
                
                display_df = detailed_df[display_cols].rename(columns={'Changes_Description': 'Changes Made'})
                
                for r in dataframe_to_rows(display_df, header=True, index=False):
                    ws.append(r)

                header_row = ws[start_row_detailed]
                for cell in header_row:
                    cell.font = Font(bold=True)
                    
                # --- Auto-fit columns ---
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            # Use date formatting if the column is 'MW Effective From' (DATE_COLUMN)
                            if DATE_COLUMN in display_df.columns:
                                date_col_index = list(display_df.columns).index(DATE_COLUMN) + 1
                                if column_letter == get_column_letter(date_col_index):
                                    if cell.value is not None and isinstance(cell.value, pd.Timestamp):
                                        text_value = cell.value.strftime('%d/%m/%Y')
                                    else:
                                        text_value = str(cell.value)
                                else:
                                    text_value = str(cell.value)
                            else:
                                text_value = str(cell.value)

                            if len(text_value) > max_length:
                                max_length = len(text_value)
                        except:
                            pass
                    adjusted_width = min(60, (max_length + 2))
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Freeze the top header rows
                ws.freeze_panes = ws['E2'] 
                
        print(f"\n--- SUCCESS ---")
        print(f"Detailed comparison report created: {output_filename}")
        
    except Exception as e:
        print(f"\nError: Failed to write to Excel. Details: {e}")

# --- Main execution block ---
if __name__ == "__main__":
    
    import glob 
    
    files_by_date, valid_files = get_file_list()
    
    if valid_files:
        file_a, file_b = get_comparison_selection(files_by_date, valid_files)
        
        if file_a and file_b:
            if file_a == file_b:
                 print("\nError: Please select two different files for comparison.")
                 df_a = load_raw_data(file_a)
                 if df_a is not None:
                    df_a_pot3 = df_a[df_a['Pot'] == 'Pot 3']
                    df_a_pot1 = df_a[df_a['Pot'] == 'Pot 1']
                    
                    print(f"\n--- Single File Analysis: {os.path.basename(file_a)} ---")
                    print(f"Pot 3 (Offshore Wind): {len(df_a_pot3):,} projects, {df_a_pot3[CAPACITY_COLUMN].sum():,.2f} MW")
                    print(f"Pot 1 (Onshore/Solar): {len(df_a_pot1):,} projects, {df_a_pot1[CAPACITY_COLUMN].sum():,.2f} MW")
            
            else:
                print(f"\nProcessing Version A: {os.path.basename(file_a)}")
                print(f"Processing Version B: {os.path.basename(file_b)}")

                df_a = load_raw_data(file_a)
                df_b = load_raw_data(file_b)

                if df_a is not None and df_b is not None:
                    df_change_summary, df_final_report = get_project_level_comparison(df_a, df_b)
                    
                    write_to_excel(df_change_summary, df_final_report, file_a, file_b)